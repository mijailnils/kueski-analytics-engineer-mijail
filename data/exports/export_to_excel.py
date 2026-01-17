# -*- coding: utf-8 -*-
"""
Export dbt models to Excel files for analysis
MANTIENE COMPATIBILIDAD CON TABLEAU DASHBOARD
"""
import duckdb
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Paths
DB_PATH = Path("dbt/kueski_finance.duckdb")
EXPORT_DIR = Path("data/exports")
EXPORT_DIR.mkdir(exist_ok=True, parents=True)

def clean_dataframe_for_excel(df):
    """Remove timezone info from datetime columns"""
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)
    return df

def format_excel(filepath):
    """Apply basic formatting to Excel file"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    wb.save(filepath)

def export_loans_for_tableau(conn):
    """Export loans con nombres de columnas para Tableau"""
    print(f"\n{'='*70}")
    print(f"Exporting: loans.xlsx (Tableau compatible)")
    print(f"{'='*70}")
    
    # Query con nombres compatibles con Tableau
    query = """
    SELECT 
        loan_id,
        user_id,
        vintage_date,
        vintage_month,
        vintage_year,
        loan_amount AS "Total Loan Amount",
        loan_term,
        interest_rate,
        loan_sequence_number,
        
        -- Flags (mantener nombres originales)
        flg_first_loan_customer,
        flg_recurrent_customer,
        
        -- También agregar versiones calculadas para Tableau
        flg_first_loan_customer AS "First Loans Count",
        CASE WHEN flg_first_loan_customer = 0 THEN 1 ELSE 0 END AS "Repeat Loans Count",
        
        -- Customer attributes
        risk_segment_customer,
        risk_band_production_customer,
        channel_customer,
        state_customer,
        city_customer,
        
        -- Funding rate components (NUEVO)
        tiie_28_rate,
        spread_rate,
        funding_cost_annual,
        funding_cost_monthly,
        funding_cost_daily,
        
        -- Revenue components (Tableau format + lowercase)
        revenue AS "Revenue Total",
        revenue,
        interest_revenue AS "Interest Total",
        interest_revenue,
        fee_revenue AS "Fees Total",
        fee_revenue,
        penalty_revenue AS "Penalties Total",
        penalty_revenue,
        
        -- Costs (Tableau format + lowercase)
        funding_cost AS "Funding Cost Total",
        funding_cost,
        credit_loss AS "Credit Loss Total",
        credit_loss,
        cogs AS "Cogs Total",
        cogs,
        cac AS "Cac Total",
        cac,
        
        -- Margins (Tableau format + lowercase)
        financial_margin AS "Financial Margin Total",
        financial_margin,
        contribution_margin AS "Contribution Margin Total",
        contribution_margin,
        net_profit AS "Net Profit Total",
        net_profit,
        
        -- Recovery
        principal_repaid,
        recovery_rate,
        
        -- Delinquency
        delinquency_status,
        is_delinquent,
        dpd_bucket,
        capital_balance,
        
        -- Delinquency Rate (calculated)
        CASE WHEN is_delinquent = 1 THEN 1.0 ELSE 0.0 END AS "Delinquency Rate",
        
        -- Metrics for aggregation
        1 AS total_loans,
        1 AS total_customers
        
    FROM main.int_loan_financials
    """
    
    df = conn.execute(query).df()
    df = clean_dataframe_for_excel(df)
    
    print(f"   Columns ({len(df.columns)} total):")
    critical_cols = [
        'flg_first_loan_customer', 'flg_recurrent_customer',
        'revenue', 'funding_cost', 'credit_loss', 'financial_margin', 
        'contribution_margin', 'net_profit', 'recovery_rate'
    ]
    for col in critical_cols:
        if col in df.columns:
            print(f"      ✅ {col}")
        else:
            print(f"      ❌ {col} MISSING!")
    
    output_path = EXPORT_DIR / 'loans.xlsx'
    df.to_excel(output_path, index=False, engine='openpyxl')
    format_excel(output_path)
    
    print(f"   ✅ Exported {len(df):,} rows")
    print(f"   File: {output_path}")
    
    return df

def export_table(conn, table_name, output_filename):
    """Export a table to Excel with formatting"""
    print(f"\n{'='*70}")
    print(f"Exporting: {table_name}")
    print(f"{'='*70}")
    
    query = f"SELECT * FROM main.{table_name}"
    df = conn.execute(query).df()
    df = clean_dataframe_for_excel(df)
    
    print(f"   Columns: {len(df.columns)}")
    
    output_path = EXPORT_DIR / output_filename
    df.to_excel(output_path, index=False, engine='openpyxl')
    format_excel(output_path)
    
    print(f"   ✅ Exported {len(df):,} rows")
    print(f"   File: {output_path}")
    
    return df

def main():
    """Export all required tables"""
    print("\n" + "="*70)
    print("KUESKI DATA EXPORT TO EXCEL")
    print("="*70)
    print(f"Database: {DB_PATH}")
    print(f"Export directory: {EXPORT_DIR}")
    
    if not DB_PATH.exists():
        print(f"\n❌ ERROR: Database not found at {DB_PATH}")
        print("Please run 'dbt run' first to create the database.")
        return
    
    # Connect to DuckDB
    conn = duckdb.connect(str(DB_PATH), read_only=True)
    
    # 1. Export loans (TABLEAU COMPATIBLE)
    df_loans = export_loans_for_tableau(conn)
    
    # 2. Export customers
    df_customers = export_table(conn, 'stg_customers', 'customers.xlsx')
    
    # 3. Export aggregated performance
    df_agg = export_table(conn, 'fct_agg_performance', 'fct_agg_performance.xlsx')
    
    # 4. Export vintage curves
    df_vintage = export_table(conn, 'mart_vintage_curves', 'vintage_curves.xlsx')
    
    # 5. Export funding cost rates (NUEVO)
    df_funding = export_table(conn, 'stg_funding_cost_rates', 'funding_cost_rates.xlsx')
    
    # Summary
    print("\n" + "="*70)
    print("EXPORT SUMMARY")
    print("="*70)
    print(f"✅ loans.xlsx")
    print(f"   - {len(df_loans):,} loans")
    print(f"   - {len(df_loans.columns)} columns")
    print(f"   - Includes: flg_first_loan_customer, flg_recurrent_customer")
    print(f"   - NEW: funding_cost, credit_loss separated")
    print()
    print(f"✅ customers.xlsx - {len(df_customers):,} rows")
    print(f"✅ fct_agg_performance.xlsx - {len(df_agg):,} rows")
    print(f"✅ vintage_curves.xlsx - {len(df_vintage):,} rows")
    print(f"✅ funding_cost_rates.xlsx - {len(df_funding):,} rows (NEW)")
    print()
    print("="*70)
    print(f"All files saved to: {EXPORT_DIR.absolute()}")
    print("="*70)
    
    conn.close()

if __name__ == "__main__":
    main()
