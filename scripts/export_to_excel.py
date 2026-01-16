# -*- coding: utf-8 -*-
"""
Export dbt models to Excel files for analysis
"""
import duckdb
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Paths - Apunta a la DB correcta
DB_PATH = Path("dbt/kueski_finance.duckdb")
EXPORT_DIR = Path("data/exports")
EXPORT_DIR.mkdir(exist_ok=True, parents=True)

def clean_dataframe_for_excel(df):
    """Remove timezone info from datetime columns"""
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)
    return df

def format_excel(filepath):
    """Apply basic formatting to Excel file"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    wb.save(filepath)
    print(f"   OK Formatted successfully")

def export_table(conn, table_name, output_filename):
    """Export a table to Excel with formatting"""
    print(f"\n{'='*70}")
    print(f"Exporting: {table_name}")
    print(f"{'='*70}")
    
    query = f"SELECT * FROM main.{table_name}"
    df = conn.execute(query).df()
    
    # Clean dataframe for Excel (remove timezone info)
    df = clean_dataframe_for_excel(df)
    
    # Show sample columns
    print(f"   Columns ({len(df.columns)} total):")
    for i, col in enumerate(df.columns[:10], 1):
        print(f"      {i}. {col}")
    if len(df.columns) > 10:
        print(f"      ... and {len(df.columns) - 10} more")
    
    output_path = EXPORT_DIR / output_filename
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    format_excel(output_path)
    
    print(f"   OK Exported {len(df):,} rows")
    print(f"   File: {output_path}")
    
    return df

def main():
    """Export all required tables"""
    print("\n" + "="*70)
    print("KUESKI DATA EXPORT TO EXCEL")
    print("="*70)
    print(f"Database: {DB_PATH}")
    print(f"Export directory: {EXPORT_DIR}")
    
    if not DB_PATH.exists():
        print(f"\nERROR: Database not found at {DB_PATH}")
        print("Please run 'dbt run' first to create the database.")
        return
    
    # Connect to DuckDB
    conn = duckdb.connect(str(DB_PATH), read_only=True)
    
    # List available tables
    tables = conn.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'main' ORDER BY table_name").fetchall()
    print(f"\nAvailable tables: {len(tables)}")
    for table in tables[:5]:
        print(f"  - {table[0]}")
    if len(tables) > 5:
        print(f"  ... and {len(tables) - 5} more")
    
    # 1. Export customers with flg_recurrent_customer
    df_customers = export_table(conn, 'customers_export', 'customers.xlsx')
    
    # 2. Export loans (int_loan_financials)
    df_loans = export_table(conn, 'int_loan_financials', 'loans.xlsx')
    
    # 3. Export aggregated performance
    df_agg = export_table(conn, 'fct_agg_performance', 'fct_agg_performance.xlsx')
    
    # 4. Export vintage curves
    df_vintage = export_table(conn, 'mart_vintage_curves', 'vintage_curves.xlsx')
    
    # Summary
    print("\n" + "="*70)
    print("EXPORT SUMMARY")
    print("="*70)
    print(f"OK customers.xlsx")
    print(f"   - {len(df_customers):,} customers")
    print(f"   - Key fields: user_id, channel, risk_segment, flg_recurrent_customer")
    print()
    print(f"OK loans.xlsx")
    print(f"   - {len(df_loans):,} loans")
    print(f"   - Key fields: loan_id, loan_sequence_number, flg_first_loan_customer,")
    print(f"                 channel_customer, risk_segment_customer")
    print()
    print(f"OK fct_agg_performance.xlsx")
    print(f"   - {len(df_agg):,} aggregated rows")
    print(f"   - Aggregated by: month, channel_customer, risk_segment_customer,")
    print(f"                    flg_recurrent_customer")
    print()
    print(f"OK vintage_curves.xlsx")
    print(f"   - {len(df_vintage):,} data points")
    print(f"   - Cohort performance over time by vintage_month, risk_segment_customer,")
    print(f"                    months_since_origination")
    print()
    print("="*70)
    print(f"All files saved to: {EXPORT_DIR.absolute()}")
    print("="*70)
    
    conn.close()

if __name__ == "__main__":
    main()
